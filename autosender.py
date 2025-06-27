import openpyxl
import smtplib
from email.mime.text import MIMEText

# Путь к файлу Excel
excel_path = r"C:\Users\vbfun\Downloads\contacts.xlsx"

# Загружаем книгу и лист
wb = openpyxl.load_workbook(excel_path)
sheet = wb[wb.sheetnames[0]]  # берем первый лист

# Читаем email из колонки J (индекс 10, J - 10-я буква, openpyxl 1-индексация)
emails = []
for row in range(2, sheet.max_row + 1):  # начиная со второй строки, если первая — заголовок
    cell = sheet.cell(row=row, column=10).value
    if cell and '@' in str(cell):
        emails.append(cell.strip())

print(f"Найдено {len(emails)} email адресов")

# Настройки почты
smtp_server = "smtp.gmail.com"
smtp_port = 587
sender_email = "titalcompanyvolodymyrbobyliev@gmail.com"  # замени на свой
app_password = "umwd ujgk hpir ykcm"       # пароль приложения Gmail

# Текст письма
subject = "TESTED IN WAR!!! Production of Ukrainian plant of fire fighting vehicles"
body = """
Dear colleges, 
We are TITAL - the Ukrainian plant of fire fighting vehicles and now we are open for cooperation. High European standards of our production are coupled with the correct Ukrainian price. We are wating for your orders! Think over your new partner - us
Head of Export Department
Volodymyr Bobyliev"""

# Создаем SMTP-сессию
server = smtplib.SMTP(smtp_server, smtp_port)
server.starttls()
server.login(sender_email, app_password)

for recipient in emails:
    msg = MIMEText(body, "plain", "utf-8")
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = recipient

    try:
        server.sendmail(sender_email, recipient, msg.as_string())
        print(f"Отправлено: {recipient}")
    except Exception as e:
        print(f"Ошибка отправки на {recipient}: {e}")

server.quit()